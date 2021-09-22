##############################################################
##
##  Daniel Jimenez, Metro, 3, 10, 2016 
##  Script to calculate the RMSE by using count and data results
##  Talk to bill about the project
##
##############################################################

from openpyxl import load_workbook
from openpyxl import Workbook
import easygui as eg
import os


##Class intended to aggregate all highway and arterial data into seperate lists
##Used to calculate rmse results
class rmse(object):
    def __init__(self):
        self.highway = []
        self.high_counts = []
        self.arterials = []
        self.art_counts = []
        
    
    def get_highway(self, value):
        self.highway.append(int(round(value,0)))

    def get_high_counts(self, value):
        self.high_counts.append(int(round(value,0)))
        
    def get_arterials(self, value):
        self.arterials.append(int(round(value,0)))
   
    def get_art_counts(self, value):
        self.art_counts.append(int(round(value,0)))
    
    def summary(self):
        hmc = sum(self.highway)
        hn = len(self.highway)
        hcounts = sum(self.high_counts)
        hrmse = self.calc_rmse(hmc,hn,hcounts)
        print "\n\nHighway Summary:"
        print "M-C^2 = {}".format(hmc)
        print "N = {}".format(hn)
        print "Sum(counts) = {}".format(hcounts)
        print "% RMSE = {}".format(hrmse)
        
        amc = sum(self.arterials)
        an = len(self.arterials)
        acounts = sum(self.art_counts)
        armse = self.calc_rmse(amc,an,acounts)
        print "\n\nArtirials Summary:"
        print "M-C^2 = {}".format(amc)
        print "N = {}".format(an)
        print "Sum(counts) = {}".format(acounts)
        print "% RMSE = {}".format(armse)

        # print self.highway
        # print self.high_counts
        # for x,y in zip(*self.highway):
            # print x,y
        
    def calc_rmse(self, mc, n, counts):
        return round((((mc/(n-1))**0.5)*100)/(counts/n),1)



#Class intended to iterate through sheet and take counts and pm2 values
#Parameters are combination on row and key columns.
#In case  columns change, its just a metter of changin the correspodning column variable to reflect desired column.         
class Praser(object):
    def __init__(self, **kwargs):
        self.data = kwargs.get('data',0)
        self.start_row = kwargs.get("start_row",0) 
        self.end_row = kwargs.get("end_row",0)
        self.freeway_col = kwargs.get("freeway_col",0)
        self.ne_counts = kwargs.get("ne_counts",0) 
        self.ne_pm2 = kwargs.get("ne_pm2" ,0)
        self.sw_counts = kwargs.get("sw_counts",0) 
        self.sw_pm2 = kwargs.get("sw_pm2",0) 
        self.sheet_name = kwargs.get("sheet_name",0)
        self.cutlines = kwargs.get("cutlines",0)
        self.start()
        
    #Driver
    def start(self):
        self.rmse = rmse()
        self.run_praser()
        
        
    #Iterates through sheet. Notice rows. Indicates that it will iterate column A from row start to row end_row
    #Avoids unnsesesary iteration and passes valid cells to check_row method
    #To be a valid cell, column A cell needs start with a number and not be null.
    def run_praser(self):
        rows = "{cutlines}{}:{cutlines}{}".format(self.start_row, self.end_row, cutlines = self.cutlines)
        for row in self.data.iter_rows(rows):
            for cell in row:
                if cell.value != None:                    
                    self.check_row(cell)

    #Identify if is a freeway row or a arterial row.
    #Passes value to rmse lists depedning on the binary freeway value    
    def check_freeway(self, row, counts, pm2):
        cell = "{}{}".format(self.freeway_col, row)
        #If cell isn't null
        if self.data[cell].internal_value:
            #Special case to sum the 111th free way hov links.
            #If link starts with 111 it adds the next row PM2 value. it only adds the north direction becasue it checks for not none values.
            if int(self.data["{}{}".format(self.cutlines,row)].internal_value.encode('utf8')[0:3]) == 111 and self.data["{}{}".format(pm2,row + 1)].internal_value is not None:
                usi5  = self.data["{}{}".format(pm2,row)].internal_value + self.data["{}{}".format(pm2,row + 1)].internal_value
                self.rmse.get_highway((usi5 - self.data["{}{}".format(counts,row)].internal_value) ** 2)
                self.rmse.get_high_counts(self.data["{}{}".format(counts,row)].internal_value)
            else:
                self.rmse.get_highway((self.data["{}{}".format(pm2,row)].internal_value - self.data["{}{}".format(counts,row)].internal_value) ** 2)
                self.rmse.get_high_counts(self.data["{}{}".format(counts,row)].internal_value)
        else:
            self.rmse.get_arterials((self.data["{}{}".format(pm2,row)].internal_value - self.data["{}{}".format(counts,row)].internal_value) ** 2)
            self.rmse.get_art_counts(self.data["{}{}".format(counts,row)].internal_value)
        
    #Check if either count of pm2 coulms are greater than 0. Avoids invalid rows that would throw off the calculation by coutning 0s
    #Internal_value method use to get the data instead of the formula. This was nessesary in order to get the formula numbers.
    def check_counts(self, row):
        ne_c = "{}{}".format(self.ne_counts,row)
        ne_pm2 = "{}{}".format(self.ne_pm2,row)
        sw_c = "{}{}".format(self.sw_counts,row)
        sw_pm2 = "{}{}".format(self.sw_pm2,row)
        
        # print ne_c, self.data[ne_c].internal_value, ne_pm2,  self.data[ne_pm2].internal_value
        
        if isinstance(self.data[ne_c].internal_value, (int,long,float)) and isinstance(self.data[ne_pm2].internal_value, (int,long,float)):            
            if self.data[ne_c].internal_value > 0 and self.data[ne_pm2].internal_value > 0:
                self.check_freeway(row, self.ne_counts, self.ne_pm2)
        if isinstance(self.data[sw_c].internal_value, (int,long, float)) and isinstance(self.data[sw_pm2].internal_value, (int,long,float)):
            if self.data[sw_c].internal_value > 0 and self.data[sw_pm2].internal_value > 0:
                self.check_freeway(row, self.sw_counts, self.sw_pm2)

    #Checks if columns is valid            
    def check_row(self, cell):
        if cell.internal_value.encode('utf8')[0:3].isdigit():            
            self.check_counts(cell.row)

 
def open_file(message):
    excel = eg.fileopenbox(msg = message, title=None, default=None)
    return excel        


    
def main():
    #Setting variables    
    #Modify as nessesary
    start_row = 3
    end_row = 139
    freeway_col = "A"
    ne_counts = "O"
    ne_pm2 = "N"
    sw_counts = "U"
    sw_pm2 = "T"
    sheet_name = "Tier 1 cutlines - PM2"
    cutlines  = "B"

    des_excel = open_file("Select Excel file where data will be copied to")
    des_book = load_workbook(des_excel)
    data = des_book.get_sheet_by_name(sheet_name)


    my_Praser = Praser(data = data, start_row = start_row, end_row = end_row, freeway_col = freeway_col, ne_counts = ne_counts, 
                        ne_pm2 = ne_pm2, sw_counts = sw_counts, sw_pm2 = sw_pm2, sheet_name = sheet_name, cutlines = cutlines)
                            
    my_Praser.rmse.summary()


if __name__ == "__main__":
    main()          
        
    

                    
                 
