##############################################################
##
##  Daniel Jimenez, Metro, 4, 8, 2016 
##  Script that put summaries reports to a result template
##
##############################################################


import openpyxl
from string import ascii_uppercase
from openpyxl.styles import Alignment, Font
from openpyxl import load_workbook
import sys
import shutil
import os
import easygui as eg

#Check if the file exist. If it doesnt it copy the template to the location
def check_file_exist():
    yes = eg.ynbox(msg="Do you have an Excel results file?", title="")
    if yes:
         return eg.fileopenbox(msg="Select results Excel template", title=None)
    else:
        path = eg.diropenbox(msg="Select location of new file", title=None)
        temp_name = eg.enterbox(msg ="Enter name of Excel results new template ", title =' ', default ='')
        fname = os.path.join(path, "{}.xlsx".format(temp_name))
        summary_source = "V:/tbm/kate/programs_v1.0/src/py/automation_files/comparison_template.xlsx"
        if not os.path.isfile(fname):
            shutil.copyfile(summary_source, fname)
        return fname

#Finds the next empty cell to put new results
#Takes in start row and col to start chekcing emtpy spaces.        
def find_column(sheet, col=0, row=0, spaces=1, **kargs):
    while True:
        if sheet.cell(row = row , column = col).value is None:  
            if sheet.cell(row = row + spaces, column = col + spaces).value is None:
                return col + 1
                break                
        col +=1

def load_book(path):
    return load_workbook(path, data_only = True)
    

def write_data(des_seet, source_sheet, source_range, des_col, format=False):
    for row in source_sheet.iter_rows(source_range):
        for cell in row:
            if format:
                des_seet.cell(row = cell.row, column = des_col).number_format = format
            des_seet.cell(row = cell.row, column = des_col).value = cell.value 
            des_seet.cell(row = cell.row, column = des_col).font = Font(name='Calibri',size=8)
            
            if cell.value in ["PM", "No service", "MD"]:    
                des_seet.cell(row = cell.row, column = des_col).alignment = Alignment(horizontal='center') 

#functions to write to different sheets 
##Uses **kargs to pass dic like parameter to different functions
#Tranfer Mode choice summary data  
def write_mcs(des,origin,**kargs):
    #Get the sheet of both origin and destinaion excel files
    ds,oris = get_sheets(des,origin,name="Mode Choice Summary")
    #Get the next avaiable cell to write new data
    des_col = find_column(ds, **kargs)
    #Call function to copy cell values from the original 
    write_data(ds, oris,"C1:C150",des_col,format="#,###")
    write_data(ds, oris,"D1:D150", des_col + 1, format = "0.00%")

    #Merges B,C cells and Aligned the title cell to the middle
    ds.merge_cells(start_row = 2, start_column = des_col, end_row=2, end_column = des_col+1)
    ds.cell(row = 2, column = des_col).alignment = Alignment(horizontal='center')
    ds.merge_cells(start_row = 3, start_column = des_col, end_row=3, end_column = des_col+1)    
    ds.cell(row = 3, column = des_col).alignment = Alignment(horizontal='center')                  
    ds.cell(row = 4, column = des_col).alignment = Alignment(horizontal='center')                  
    ds.cell(row = 4, column = des_col + 1).alignment = Alignment(horizontal='center')                  

#transit boardings summary data  
def write_tb(des,origin,**kargs):
    ds,oris = get_sheets(des,origin,name="Transit Boardings")
    des_col = find_column(ds, **kargs)
    write_data(ds, oris,"A1:A350",des_col)
    write_data(ds, oris,"B1:B350",des_col + 1,format = "#,###")
    write_data(ds, oris,"C1:C350",des_col + 2,format = "#,###")

    ds.merge_cells(start_row = 1, start_column = des_col, end_row=1, end_column = des_col+2)
    ds.cell(row = 1, column = des_col).alignment = Alignment(horizontal='center')
    ds.merge_cells(start_row = 2, start_column = des_col, end_row=2, end_column = des_col+2)    
    ds.cell(row = 2, column = des_col).alignment = Alignment(horizontal='center')          

#average travel length summary data    
def write_atl(des,origin,**kargs):
    ds,oris = get_sheets(des,origin,name="Average Trip Length")
    des_col = find_column(ds, **kargs)
    write_data(ds, oris,"C1:C35",des_col)

    ds.cell(row = 1, column = des_col).alignment = Alignment(horizontal='center')

#Returns letters from alphabet string
#Creates a list of two alphabets sets to accounts for columns larger than z.     
def get_d2d_columns(letter):
    letters = ascii_uppercase
    alphabet = list(letters) + ["A{}".format(i) for i in letters]
    s_col = alphabet.index(letter)  
    end_col_index = s_col + 4  
    return alphabet[s_col:end_col_index]

    
def writed2d(des,origin,**kargs):
    letters = get_d2d_columns(kargs["s_col"])
    ds,oris = get_sheets(des,origin,alt_name="dist2dist", **kargs)
    des_col = find_column(ds, **kargs)
    write_data(ds, oris,"{c}1:{c}60".format(c = letters[0]),des_col, format = "0.00%")
    write_data(ds, oris,"{c}1:{c}60".format(c = letters[1]),des_col + 1, format = "0.00%")
    write_data(ds, oris,"{c}1:{c}60".format(c = letters[2]),des_col + 2, format = "0.00%")
    write_data(ds, oris,"{c}1:{c}60".format(c = letters[3]),des_col + 3, format = "0.00%")
    ds.merge_cells(start_row = 1, start_column = des_col, end_row=1, end_column = des_col+3)
    ds.cell(row = 1, column = des_col).alignment = Alignment(horizontal='center')  
    ds.cell(row = 1, column = des_col).value =  get_title(origin) 
            
def transfer_data(des_book,origin_book):
    write_mcs(des_book,origin_book,col=2,row=7)
    write_tb(des_book,origin_book,col=1,row=3)
    write_atl(des_book,origin_book,col=2,row=3)
    writed2d(des_book,origin_book,col=2,row=4, name="dist2dist_HBW", s_col = "C")
    writed2d(des_book,origin_book,col=2,row=4, name="dist2dist_HBNW", s_col = "J")
    writed2d(des_book,origin_book,col=2,row=4, name="dist2dist_NHB", s_col = "Q")
    writed2d(des_book,origin_book,col=2,row=4, name="dist2dist_All", s_col = "X")

#Use to get title for d2d sheets that do not have a title
#It grabs title from the Average trip lengh source sheet
#If it does not exits, it will return a no title string    
def get_title(source_book):
    try:
        sheet =   source_book.get_sheet_by_name("Average Trip Length")
        return sheet["C1"].value
    except:
        return "No Title Found"    

#Needed to get the sheet by name
#Alt_name used to get a different name in case the source and destinaion sheet names do not mach
#Alt_name Used in the d2d sheets        
def get_sheets(des,origin,name="", alt_name= False, **kargs):
    if alt_name:
        origin_sheet = origin.get_sheet_by_name(alt_name)
    else:
        origin_sheet = origin.get_sheet_by_name(name)
    des_sheet = des.get_sheet_by_name(name)
    return des_sheet,origin_sheet    

def main():
    result_path = check_file_exist()
    result_excel = load_book(result_path)    
    
    while True:
        source_path = eg.fileopenbox(title="Select results Excel template", msg="When done just close the dialog or click cancel")
        if source_path is not None:
            source_excel = load_book(source_path)
            transfer_data(result_excel,source_excel)  
            result_excel.save(result_path) 
        else:    
            break
 
    
if __name__ == "__main__":    
    main()