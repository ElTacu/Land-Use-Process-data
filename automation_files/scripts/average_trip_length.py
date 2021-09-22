##############################################################
##
##  AverageTripLength.py
##
##  --Script that reads the 621 summary and reports lines by mode.
##
##  Daniel Jimenez, Metro, March 15, 2016
##
##############################################################


import openpyxl
from openpyxl import load_workbook
from collections import defaultdict
import sys
import subprocess
import os

def avgTripLength(sheet):
    sheet["C1"].value = sys.argv[2]
    avg_trips = {}
    #os.path.dirname(sys.argv[1] = ../model/reports
    with open(os.path.join(os.path.dirname(sys.argv[1]),"auto_files/average_trip_length.csv")) as f:
        for i in f:
             p,k,v,_ = i.split(",")
             avg_trips[k] = float(v)

    for row in sheet.iter_rows("B3:B17"):
        for cell in row:
            sheet["C{}".format(cell.row)].value = avg_trips.get(cell.value,"")
    sheet["C19"].value = avg_trips["College"]
    sheet["C21"].value = avg_trips["All Purpose"]

def read_data(file):
    with open(file) as t:
       di = {}
       for line in t:
            from_, to, _, value, _ = line.split(",")
            di["{}-{}".format(from_, to)] = value
    return di        

#Populates the d2d section of sheet    
def dit2dit(sheet):
    dist = defaultdict(list)
    #os.path.dirname(sys.argv[1] = ../model/reports
    p_trips = read_data(os.path.join(os.path.dirname(sys.argv[1]),"auto_files/person_trips.csv"))
    t_length = read_data(os.path.join(os.path.dirname(sys.argv[1]),"auto_files/total_trip_lengths.csv"))
    
    for k in p_trips.keys():
        avg = float(t_length[k]) / float(p_trips[k]) 
        district = int(k.split("-")[0])
        dist[district].append(avg)
        
    #25 in format to reference row 25 in sheet
    for i in range(1,7):
        sheet["C{}".format(25 + i)].value = round(sum(dist[i])/len(dist[i]),1)

        
def run_r(root):
    subprocess.call("Rscript --slave --vanilla V:/tbm/kate/programs_v1.0/src/py/automation_files/scripts/tripLengths.R {}".format(root))
    subprocess.call("Rscript --slave --vanilla V:/tbm/kate/programs_v1.0/src/py/automation_files/scripts/avetriplenght.R {}".format(root))


root = os.path.dirname(os.path.dirname(os.path.dirname(sys.argv[1])))
run_r(root)    
excel_file = sys.argv[1]
des_excel = excel_file
des_book = load_workbook(des_excel)
mode_sheet = des_book.get_sheet_by_name("Average Trip Length")
avgTripLength(mode_sheet)
dit2dit(mode_sheet)
des_book.save(des_excel)   