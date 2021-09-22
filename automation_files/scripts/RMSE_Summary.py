##############################################################
##
##  Daniel Jimenez, Metro, 3, 10, 2016 
##  Script to calculate the RMSE by using count and data results
##
##############################################################

from RMSE_Auto import Praser
import openpyxl
from openpyxl import load_workbook
import sys
import os

def write_RMSE(sheet,rmse,col,row):
    sheet["{}{}".format(col,row)].value = sum(rmse.highway)
    sheet["{}{}".format(col,row + 1)].value = len(rmse.highway)
    sheet["{}{}".format(col,row + 2)].value = sum(rmse.high_counts)
    sheet["{}{}".format(col,row + 3)].value = rmse.calc_rmse(sum(rmse.highway),len(rmse.highway),sum(rmse.high_counts))
    
    sheet["{}{}".format(col,row + 6)].value = sum(rmse.arterials)
    sheet["{}{}".format(col,row + 7)].value = len(rmse.arterials)
    sheet["{}{}".format(col,row + 8)].value = sum(rmse.art_counts)
    sheet["{}{}".format(col,row + 9)].value = rmse.calc_rmse(sum(rmse.arterials),len(rmse.arterials),sum(rmse.art_counts))
    

#Setting variables    
#Modify as nessesary
start_row = 3
end_row = 152
freeway_col = "A"
ne_counts = "O"
ne_pm2 = "N"
sw_counts = "U"
sw_pm2 = "T"
cutlines  = "B"


source_excel = sys.argv[1]
source_book = load_workbook(source_excel, data_only = True)
sheet_name = "Tier 1 cutlines - PM2"
source_data = source_book.get_sheet_by_name(sheet_name)
source_book.save(source_excel)


my_praser_pm = Praser(data = source_data, start_row = start_row, end_row = end_row, freeway_col = freeway_col, ne_counts = ne_counts, 
                    ne_pm2 = ne_pm2, sw_counts = sw_counts, sw_pm2 = sw_pm2, sheet_name = sheet_name, cutlines = cutlines)

sheet_name_awd = "Tier 1 cutlines - AWD"
source_data_awd = source_book.get_sheet_by_name(sheet_name_awd)
my_praser_awd = Praser(data = source_data_awd, start_row = start_row, end_row = end_row, freeway_col = freeway_col, ne_counts = ne_counts, 
                    ne_pm2 = ne_pm2, sw_counts = sw_counts, sw_pm2 = sw_pm2, sheet_name = sheet_name_awd, cutlines = cutlines)


#write data to summary file
excel_file = sys.argv[1]
                    
des_excel = excel_file
des_book = load_workbook(des_excel)
data = des_book.get_sheet_by_name("RMSE")

write_RMSE(data,my_praser_pm.rmse,"B",4)                    
write_RMSE(data,my_praser_awd.rmse,"C",4)                    
                    
des_book.save(des_excel)                    
                    

                    
                 
