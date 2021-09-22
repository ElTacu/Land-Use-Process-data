##############################################################
##
##
##  Daniel Jimenez, Metro, March 15, 2016
##
##############################################################

from peakSpreadAllDayVolumes import get_data, times, volumes
import openpyxl
from openpyxl import load_workbook
from Emmy_Functions import Emmy_Functions
import sys
import os
import subprocess

bank_path = os.path.join(os.path.dirname(os.path.dirname(sys.argv[1])),"peak/assignPeakSpread/emmebank")
bank =  bank_path    
#Needs to change to reflect new path
excel_file = sys.argv[1]
# excel_file = "I:/ModServStaff/jimenez/AutomatedCalibrationSummaries/CutlineComparison.xlsx"
emmy_bank = Emmy_Functions(bank)
auto_links, capacity = get_data(emmy_bank)

#Write data to the excel file
des_excel = excel_file
des_book = load_workbook(des_excel)
mode_sheet = des_book.get_sheet_by_name("emme_volumes")

def write_file(sheet,links, capacity ):
    header = ["UNIQUEID","From","To","Capacity"] + times()
    for i, filed in enumerate(header):
         sheet.cell(row = 1, column = (i + 1)).value = filed
    
    row = 2
    for k,v in links.iteritems():
        unique = "L{}".format(k.replace("-",""))
        fr, to = k.split("-")
        values = [unique, fr, to, capacity[k]] + volumes(v)
        for i, value in enumerate(values):        
            sheet.cell(row = row, column = (i + 1)).value = value
        row += 1  
write_file(mode_sheet, auto_links, capacity)
des_book.save(des_excel)


from win32com.client import Dispatch
import win32com
file_path = sys.argv[1]

try:
    wb = win32com.client.GetObject(file_path)
    wb.Save()
    wb.Close()
except:    
    subprocess.call("Rscript --slave --vanilla V:/tbm/kate/programs_v1.0/src/py/automation_files/scripts/open_excel.R")